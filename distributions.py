"""
distributions.py
================
All probability-distribution sampling functions used by the simulation.

The project brief requires that every random variate be generated from
U(0, 1) via one of the algorithms taught in class:

  Box-Muller        - Normal distributions
                      Used for: charging battery level, body-art glitter,
                                food ordering service time, and (after
                                fitting) MainStage show duration.

  Inverse Transform - Exponential, Uniform (continuous & discrete)
                      Used for: most arrival processes, SideStage show
                                duration, food prep times, henna body-art,
                                FriendsGroup size, etc.

  Composition       - Piecewise PDFs that split naturally into pieces
                      Used for: PhotoStation service duration (3 linear
                                pieces over [1, 4]).

  Accept-Reject     - Bounded PDFs without a closed-form inverse CDF
                      Used for: DJStage stay duration (piecewise PDF
                                over [20, 60] with a triangular shape).

Additionally this module exposes:
  - sample_charging_duration : power-law CDF, closed-form inverse
  - load_sample_data         : reads the two-sheet Excel file
  - fit_*                    : MLE fits for Exponential / Normal / Uniform
  - kolmogorov_smirnov_statistic : KS goodness-of-fit metric
  - best_fit                 : picks the best-fitting distribution by KS
  - fit_from_excel           : Excel data → ready-to-inject samplers

Each sampler documents the PDF / CDF / inverse-CDF (or the proposal +
acceptance test for accept-reject) in its docstring so the report can
quote the formulas directly.
"""

import math
import random
from typing import Optional, List, Tuple, Callable, Dict


# ─────────────────────────────────────────────────────────────────────────────
# 1. PRIMITIVE UNIFORM SAMPLER
# ─────────────────────────────────────────────────────────────────────────────

def sample_uniform_01() -> float:
    """Draw a single U(0, 1) sample."""

    # random.random() returns a random number between 0 and 1.
    return random.random()


# ─────────────────────────────────────────────────────────────────────────────
# 2. BOX-MULLER TRANSFORM  →  Normal distribution
# ─────────────────────────────────────────────────────────────────────────────

class BoxMuller:
    """
    Generates Normal(mu, sigma) samples via the Box-Muller transform.

    The transform converts two independent U(0,1) variates (u1, u2) into
    two independent standard-normal variates:
        Z1 = sqrt(-2 ln u1) * cos(2π u2)
        Z2 = sqrt(-2 ln u1) * sin(2π u2)
    Then X = mu + sigma * Z is returned.

    We cache the second variate Z2 to avoid wasting it.
    """

    def __init__(self):
        # This variable saves one generated normal value for later use.
        self._cached: Optional[float] = None

    def reset(self) -> None:
        """Drop the cached variate. Call between simulation runs so that
        a cached Z2 from a previous (un-seeded) run doesn't leak into a
        freshly seeded run and break reproducibility."""

        # Clear the saved value.
        self._cached = None

    def sample(self, mu: float = 0.0, sigma: float = 1.0) -> float:
        """Return one Normal(mu, sigma) sample."""

        # If we already have a saved normal value, use it now.
        if self._cached is not None:
            z = self._cached

            # Clear the cache after using it.
            self._cached = None

            # Convert the standard normal value to Normal(mu, sigma).
            return mu + sigma * z

        # Generate two uniform random numbers.
        u1 = sample_uniform_01()
        u2 = sample_uniform_01()

        # Avoid log(0), because log(0) is not defined.
        while u1 == 0.0:
            u1 = sample_uniform_01()

        # Calculate the Box-Muller magnitude.
        magnitude = math.sqrt(-2.0 * math.log(u1))

        # Create two standard normal values.
        z1 = magnitude * math.cos(2.0 * math.pi * u2)
        z2 = magnitude * math.sin(2.0 * math.pi * u2)

        # Save the second value for the next call.
        self._cached = z2

        # Return the first value converted to Normal(mu, sigma).
        return mu + sigma * z1


# Module-level singleton so all callers share the cached variate
_box_muller = BoxMuller()


def sample_normal(mu: float, sigma: float) -> float:
    """Sample from Normal(mu, sigma) using Box-Muller."""

    # Use the shared BoxMuller object.
    return _box_muller.sample(mu, sigma)


def reset_box_muller() -> None:
    """Reset the Box-Muller cache. Call at the start of each replication."""

    # Clear any saved normal value before a new simulation run.
    _box_muller.reset()


# ─────────────────────────────────────────────────────────────────────────────
# 3. INVERSE TRANSFORM  →  Exponential, Uniform
# ─────────────────────────────────────────────────────────────────────────────

def sample_exponential(mean: float) -> float:
    """
    Sample from Exponential(mean) via Inverse Transform.

    CDF: F(x) = 1 - exp(-x / mean)
    Inverse: x = -mean * ln(1 - U)  =  -mean * ln(U)   (U = 1-U is also uniform)
    """

    # Generate a uniform random number.
    u = sample_uniform_01()

    # Avoid log(0).
    while u == 0.0:
        u = sample_uniform_01()

    # Apply the inverse CDF formula.
    return -mean * math.log(u)


def sample_continuous_uniform(a: float, b: float) -> float:
    """
    Sample from Uniform(a, b) via Inverse Transform.

    CDF: F(x) = (x - a) / (b - a)
    Inverse: x = a + (b - a) * U
    """

    # Generate a uniform random number.
    u = sample_uniform_01()

    # Scale the number from [0,1] to [a,b].
    return a + (b - a) * u


def sample_discrete_uniform(a: int, b: int) -> int:
    """
    Sample from Discrete Uniform {a, a+1, ..., b} via Inverse Transform.

    Maps U(0,1) → integer in [a, b] with equal probability 1/(b-a+1).
    """

    # Generate a uniform random number.
    u = sample_uniform_01()

    # Convert the uniform number to an integer between a and b.
    return a + int(u * (b - a + 1))


def sample_poisson(lam: float) -> int:
    """
    Sample from Poisson(lambda) via Inverse Transform (sequential method).

    Used for fitting/generating arrival count data.
    """

    # This value is the stopping threshold.
    L = math.exp(-lam)

    # k counts how many random numbers were multiplied.
    k = 0

    # p is the running product of uniform random numbers.
    p = 1.0

    # Keep multiplying until p becomes small enough.
    while p > L:
        k += 1
        p *= sample_uniform_01()

    # Return the Poisson sample.
    return k - 1


# ─────────────────────────────────────────────────────────────────────────────
# 4. COMPOSITION METHOD  →  PhotoStation duration
# ─────────────────────────────────────────────────────────────────────────────
#
# PDF of photo session duration (x in minutes):
#   f(x) = x/6,         1 <= x < 2   (linear, triangular-like piece)
#   f(x) = x/5 + 1/8,   2 <= x < 3
#   f(x) = 1/8,         3 <= x < 4
#
# We verify the pdf integrates to 1:
#   ∫[1,2] x/6 dx = [x²/12]_1^2 = (4-1)/12 = 3/12 = 1/4
#   ∫[2,3] (x/5 + 1/8) dx = [x²/10 + x/8]_2^3
#                          = (9/10+3/8) - (4/10+2/8)
#                          = (9/10-4/10) + (3/8-2/8) = 1/2 + 1/8 = 5/8
#   ∫[3,4] 1/8 dx = 1/8
#   Total = 1/4 + 5/8 + 1/8 = 2/8 + 5/8 + 1/8 = 8/8 = 1  ✓
#
# Component weights: w1=1/4, w2=5/8, w3=1/8
# Conditional CDFs:
#   Piece 1 (w1=0.25): F1(x) = (x²-1)/3       → x = sqrt(1 + 3U)
#   Piece 2 (w2=0.625): normalised density = (x/5+1/8)/(5/8)
#     F2(x|2<=x<3) = (x²/10 + x/8 - (4/10+2/8)) / (5/8)
#     Solve for x via quadratic: x²/10 + x/8 - C = 0 where C = F2*5/8 + (4/10+2/8)
#   Piece 3 (w3=0.125): F3(x|3<=x<4) = (x-3)  → x = 3 + U
#

# These are the probabilities of choosing each PDF piece.
_PHOTO_W1 = 1 / 4
_PHOTO_W2 = 5 / 8
_PHOTO_W3 = 1 / 8


def _photo_piece1_inverse(u: float) -> float:
    """Inverse CDF for piece 1: x in [1,2), f(x)=x/6."""

    # Use the inverse formula for the first piece.
    return math.sqrt(1.0 + 3.0 * u)


def _photo_piece2_inverse(u: float) -> float:
    """Inverse CDF for piece 2: x in [2,3), f(x)=x/5+1/8."""

    # Calculate the constant used in the quadratic equation.
    C = (5.0 / 8.0) * u + 4.0 / 10.0 + 2.0 / 8.0

    # Coefficients of the quadratic equation.
    a_coef, b_coef, c_coef = 4.0, 5.0, -40.0 * C

    # Calculate the discriminant.
    disc = b_coef ** 2 - 4.0 * a_coef * c_coef

    # Return the positive root.
    return (-b_coef + math.sqrt(disc)) / (2.0 * a_coef)


def _photo_piece3_inverse(u: float) -> float:
    """Inverse CDF for piece 3: x in [3,4), f(x)=1/8 (uniform)."""

    # The third piece is uniform between 3 and 4.
    return 3.0 + u


def sample_photo_duration() -> float:
    """
    Sample photo session duration using the Composition method.

    Steps:
        1. Pick component with probability proportional to weights.
        2. Sample from that component's conditional distribution.
    """

    # First random number chooses the piece.
    u_comp = sample_uniform_01()

    # Second random number samples inside the chosen piece.
    u_val  = sample_uniform_01()

    # Choose piece 1.
    if u_comp < _PHOTO_W1:
        return _photo_piece1_inverse(u_val)

    # Choose piece 2.
    elif u_comp < _PHOTO_W1 + _PHOTO_W2:
        return _photo_piece2_inverse(u_val)

    # Choose piece 3.
    else:
        return _photo_piece3_inverse(u_val)


# ─────────────────────────────────────────────────────────────────────────────
# 5. ACCEPT-REJECT  →  DJStage duration
# ─────────────────────────────────────────────────────────────────────────────
#
# PDF of DJ stage stay duration (x in minutes):
#   f(x) = (x - 20) / 600,            20 <= x <= 40
#   f(x) = (60 - x) / 600 + 1/30,     40 <= x <= 50
#   f(x) = (60 - x) / 600,            50 <= x <= 60
#   f(x) = 0,                          otherwise
#
# Support: [20, 60].  We use U[20,60] as the proposal distribution g(x)=1/40.
# Maximum of f(x):
#   At x=40: f(40) = 20/600 = 1/30
#   At x=40 (from right): f(40) = 20/600 + 1/30 = 1/30 + 1/30 = 2/30 = 1/15
#   So f_max = 1/15
# Acceptance constant c = f_max / g(x) = (1/15) / (1/40) = 40/15 = 8/3
#

# DJ duration support range.
_DJ_SUPPORT_MIN = 20.0
_DJ_SUPPORT_MAX = 60.0

# Maximum PDF value.
_DJ_F_MAX       = 1.0 / 15.0

# Proposal distribution density.
_DJ_G           = 1.0 / (_DJ_SUPPORT_MAX - _DJ_SUPPORT_MIN)

# Accept-reject constant.
_DJ_C           = _DJ_F_MAX / _DJ_G


def _dj_pdf(x: float) -> float:
    """Evaluate the DJStage stay-duration PDF at x."""

    # First part of the PDF.
    if 20.0 <= x <= 40.0:
        return (x - 20.0) / 600.0

    # Second part of the PDF.
    elif 40.0 < x <= 50.0:
        return (60.0 - x) / 600.0 + 1.0 / 30.0

    # Third part of the PDF.
    elif 50.0 < x <= 60.0:
        return (60.0 - x) / 600.0

    # Outside the support, density is zero.
    else:
        return 0.0


def sample_dj_duration() -> float:
    """
    Sample DJStage stay duration using the Accept-Reject method.

    Proposal: Uniform[20, 60].
    Accept x if U <= f(x) / (c * g(x)).
    """

    # Keep trying until a value is accepted.
    while True:
        # Sample a candidate value from Uniform[20,60].
        x = sample_continuous_uniform(_DJ_SUPPORT_MIN, _DJ_SUPPORT_MAX)

        # Sample a random number for the acceptance test.
        u = sample_uniform_01()

        # Accept the candidate if it passes the test.
        if u <= _dj_pdf(x) / (_DJ_C * _DJ_G):
            return x


# ─────────────────────────────────────────────────────────────────────────────
# 6. CHARGING STATION  →  custom power-law CDF
# ─────────────────────────────────────────────────────────────────────────────
#
# Battery level b ~ Normal(40, 15)  (clamped to [0, 100))
# alpha = 100 / (100 - b)
#
# f(t) = (alpha / 40^alpha) * (40 - t)^(alpha-1),  0 <= t <= 40
#
# CDF: F(t) = 1 - ((40-t)/40)^alpha
# Inverse CDF: t = 40 * (1 - (1-U)^(1/alpha))
#

def sample_charging_duration(battery_level):
    """
    Sample charging duration given the entity's battery level.

    Args:
        battery_level: percentage in [0, 100) from Normal(40,15).
    """

    # Keep battery level inside the valid range.
    if battery_level < 0.0:
        battery_level = 0.0
    if battery_level > 99.9:
        battery_level = 99.9

    # Calculate alpha according to the battery level.
    alpha = 100.0 / (100.0 - battery_level)

    # Generate a uniform random number.
    u = sample_uniform_01()

    # Apply the inverse CDF formula.
    return 40.0 * (1.0 - (1.0 - u) ** (1.0 / alpha))


def sample_battery_level(cfg_mean=40.0, cfg_std=15.0):
    """Sample battery arrival percentage ~ Normal(mean, std), clamped to [0, 99.9]."""

    # Sample battery level from a normal distribution.
    b = sample_normal(cfg_mean, cfg_std)

    # Keep battery level inside the valid range.
    if b < 0.0:
        b = 0.0
    if b > 99.9:
        b = 99.9

    # Return the valid battery level.
    return b


# ─────────────────────────────────────────────────────────────────────────────
# 7. CONVENIENCE WRAPPERS for common distributions used in the simulation
# ─────────────────────────────────────────────────────────────────────────────

def sample_body_art_duration(art_type):
    """
    Sample body art service duration by art type.
        'glitter': Normal(15, 3) via Box-Muller
        'neon'   : Exponential(12) via Inverse Transform
        'henna'  : Uniform[17, 22] via Inverse Transform
    """

    # Glitter duration is sampled from a normal distribution.
    if art_type == 'glitter':
        duration = sample_normal(15.0, 3.0)

        # Make sure the duration is not too small.
        if duration < 1.0:
            duration = 1.0

        return duration

    # Neon duration is sampled from an exponential distribution.
    elif art_type == 'neon':
        return sample_exponential(12.0)

    # Henna duration is sampled from a continuous uniform distribution.
    elif art_type == 'henna':
        return sample_continuous_uniform(17.0, 22.0)

    # If the art type is unknown, raise an error.
    else:
        raise ValueError("Unknown art type: " + str(art_type))


def sample_food_service_time(food_service_mean=5.0, food_service_std=1.5):
    """Order/payment service time ~ Normal(5, 1.5), minimum 0.5 min."""

    # Sample service time from a normal distribution.
    duration = sample_normal(food_service_mean, food_service_std)

    # Make sure the service time is not too small.
    if duration < 0.5:
        duration = 0.5

    return duration


def sample_art_type() -> str:
    """Choose body-art type: 'glitter'(0.3), 'neon'(0.3), 'henna'(0.4)."""

    # Generate a random number to choose the art type.
    u = sample_uniform_01()

    # 30% chance for glitter.
    if u < 0.3:
        return 'glitter'

    # 30% chance for neon.
    elif u < 0.6:
        return 'neon'

    # 40% chance for henna.
    else:
        return 'henna'


def sample_food_restaurant(burger_prob: float, pizza_prob: float) -> str:
    """Choose restaurant based on configured probabilities."""

    # Generate a random number to choose the restaurant.
    u = sample_uniform_01()

    # Choose burger according to burger probability.
    if u < burger_prob:
        return 'burger'

    # Choose pizza according to pizza probability.
    elif u < burger_prob + pizza_prob:
        return 'pizza'

    # Otherwise choose Asian food.
    else:
        return 'asian'


# ─────────────────────────────────────────────────────────────────────────────
# 8. DISTRIBUTION FITTING  (for samples_for_simulation.xlsx)
# ─────────────────────────────────────────────────────────────────────────────

def fit_exponential(data: List[float]) -> float:
    """
    Fit Exponential distribution to data via MLE.
    MLE estimate: lambda_hat = 1 / x_bar  → mean = x_bar
    """

    # Make sure the dataset is not empty.
    if not data:
        raise ValueError("Empty dataset")

    # For exponential distribution, the MLE mean is the sample average.
    return sum(data) / len(data)


def fit_normal(data: List[float]) -> Tuple[float, float]:
    """
    Fit Normal distribution via MLE.
    Returns (mu, sigma).
    """

    # Get the number of observations.
    n = len(data)

    # Need at least two values to calculate standard deviation.
    if n < 2:
        raise ValueError("Need at least 2 data points")

    # Calculate the sample mean.
    mu = sum(data) / n

    # Calculate the sample standard deviation.
    sigma = math.sqrt(sum((x - mu) ** 2 for x in data) / (n - 1))

    # Return the fitted parameters.
    return mu, sigma


def fit_uniform(data: List[float]) -> Tuple[float, float]:
    """
    Fit Uniform(a, b) distribution via MLE.
    MLE: a = min(data), b = max(data).
    """

    # For uniform distribution, use the minimum and maximum values.
    return min(data), max(data)


def kolmogorov_smirnov_statistic(data: List[float],
                                  cdf_func) -> float:
    """
    Compute the KS test statistic D = max |F_n(x) - F(x)|.
    Used to evaluate goodness-of-fit for a candidate distribution.
    """

    # Sort the data from smallest to largest.
    sorted_data = sorted(data)

    # Number of observations.
    n = len(sorted_data)

    # Start with zero as the maximum difference.
    d_max = 0.0

    # Go over every observation.
    for i, x in enumerate(sorted_data):

        # Empirical CDF just above the current point.
        f_empirical_upper = (i + 1) / n

        # Empirical CDF just below the current point.
        f_empirical_lower = i / n

        # Theoretical CDF value from the tested distribution.
        f_theoretical = cdf_func(x)

        # Update the largest difference.
        d_max = max(d_max,
                    abs(f_empirical_upper - f_theoretical),
                    abs(f_empirical_lower - f_theoretical))

    # Return the KS statistic.
    return d_max


def load_sample_data(xlsx_path: str) -> dict:
    """
    Load the two sheets from samples_for_simulation.xlsx.
    Returns dict with keys 'sheet1' and 'sheet2' containing lists of values.

    NOTE: Sheet 1 is assumed to contain FriendsGroup inter-arrival times.
          Sheet 2 is assumed to contain MainStage performance durations.
    """

    try:
        # Import openpyxl only when this function is used.
        import openpyxl

        # Open the Excel file in read-only mode.
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

        # Dictionary that will store values from each sheet.
        result = {}

        # Go over all sheets in the workbook.
        for name in wb.sheetnames:
            ws = wb[name]
            values = []

            # Read all cells in the sheet.
            for row in ws.iter_rows(values_only=True):
                for cell in row:

                    # Keep only numeric values.
                    if isinstance(cell, (int, float)) and cell is not None:
                        values.append(float(cell))

            # Save the numeric values of this sheet.
            result[name] = values

        # Close the workbook.
        wb.close()

        # Return all loaded data.
        return result

    except Exception as e:
        # If the file cannot be loaded, print a warning and return empty data.
        print(f"[WARNING] Could not load Excel data: {e}")
        return {}


# ─────────────────────────────────────────────────────────────────────────────
# 9. AUTOMATIC DISTRIBUTION SELECTION  (fit Excel data → ready-to-use samplers)
# ─────────────────────────────────────────────────────────────────────────────
#
# Section 8 supplies the raw MLE fits (fit_exponential/normal/uniform) and the
# KS statistic.  This section orchestrates them: for each empirical dataset it
# fits all three candidate distributions, picks the one with the smallest KS
# statistic, and returns a callable sampler the engine can inject directly.
#
# Usage:
#     from distributions import fit_from_excel
#     samplers = fit_from_excel('samples_for_simulation.xlsx')
#     sim = Simulation(
#         cfg,
#         friends_arrival_sampler     = samplers['friends_interarrival'],
#         main_stage_duration_sampler = samplers['main_stage_duration'],
#     )
#


# ── CDF helpers for the KS goodness-of-fit test ──────────────────────────────

def _exponential_cdf(mean: float) -> Callable[[float], float]:
    # Return a function that calculates the exponential CDF.
    return lambda x: 1.0 - math.exp(-x / mean) if x >= 0 else 0.0


def _normal_cdf(mu: float, sigma: float) -> Callable[[float], float]:
    # Define the normal CDF function.
    def cdf(x: float) -> float:

        # Convert x to a standard normal value.
        z = (x - mu) / sigma

        # Use the error function to calculate the normal CDF.
        return 0.5 * (1.0 + math.erf(z / math.sqrt(2.0)))

    # Return the CDF function.
    return cdf


def _uniform_cdf(a: float, b: float) -> Callable[[float], float]:
    # Return a function that calculates the uniform CDF.
    return lambda x: max(0.0, min(1.0, (x - a) / (b - a)))


# ── Fit candidate distributions and select best via KS statistic ─────────────

def best_fit(data: List[float], label: str = '',
             alpha: float = 0.01) -> Tuple[str, dict, Callable]:
    """
    Try Exponential, Normal, and Uniform fits on the data.
    Select the distribution with the smallest KS statistic, then run a
    formal Kolmogorov-Smirnov test on the winner at significance `alpha`.

    The critical values come from the course formula sheet (Table 1) and
    use the Lilliefors-style adjusted statistic when the fitted
    distribution had its parameters estimated from the data:
        Normal(mu_hat, sigma_hat): adj = (sqrt(n) - 0.01 + 0.85/sqrt(n)) * D
        Exponential(mean_hat):     adj = (sqrt(n) + 0.26 + 0.5/sqrt(n)) * (D - 0.2/n)
        All other / params known:  adj = (sqrt(n) + 0.12 + 0.11/sqrt(n)) * D

    Returns:
        (dist_name, params_dict, sampler_callable)
    """

    # Make sure the dataset is not empty.
    if not data:
        raise ValueError(f"Empty data for '{label}'")

    # Dictionary that stores all candidate distributions.
    candidates = {}

    # Exponential

    # Fit exponential distribution.
    mean_exp = fit_exponential(data)

    # Calculate KS statistic for exponential fit.
    ks_exp   = kolmogorov_smirnov_statistic(data, _exponential_cdf(mean_exp))

    # Save exponential candidate.
    candidates['Exponential'] = {
        'ks': ks_exp,
        'params': {'mean': mean_exp},
        'sampler': lambda m=mean_exp: sample_exponential(m),
    }

    # Normal

    # Fit normal distribution.
    mu_n, sigma_n = fit_normal(data)

    # Calculate KS statistic for normal fit.
    ks_norm       = kolmogorov_smirnov_statistic(data, _normal_cdf(mu_n, sigma_n))

    # Save normal candidate.
    candidates['Normal'] = {
        'ks': ks_norm,
        'params': {'mu': mu_n, 'sigma': sigma_n},
        'sampler': lambda m=mu_n, s=sigma_n: max(0.1, sample_normal(m, s)),
    }

    # Uniform

    # Fit uniform distribution.
    a_u, b_u = fit_uniform(data)

    # Calculate KS statistic for uniform fit.
    ks_uni   = kolmogorov_smirnov_statistic(data, _uniform_cdf(a_u, b_u))

    # Save uniform candidate.
    candidates['Uniform'] = {
        'ks': ks_uni,
        'params': {'a': a_u, 'b': b_u},
        'sampler': lambda a=a_u, b=b_u: sample_continuous_uniform(a, b),
    }

    # Choose best

    # Select the distribution with the smallest KS statistic.
    best_name = min(candidates, key=lambda k: candidates[k]['ks'])

    # Get the selected distribution data.
    best      = candidates[best_name]

    # Print fitting results for the user.
    print(f"\n[Distribution Fitting] '{label}'")
    print(f"  n = {len(data)}")

    # Print each candidate and mark the selected one.
    for name, info in candidates.items():
        mark = ' ◄ SELECTED' if name == best_name else ''
        print(f"  {name:12s}: KS={info['ks']:.4f}  params={info['params']}{mark}")

    # Formal KS test on the selected distribution.
    # Critical-value table (course formula sheet, page 2).
    # Rows = test variant; columns = alpha.
    ks_table = {
        'Normal_lilliefors': {0.15: 0.775, 0.10: 0.819, 0.05: 0.895,
                              0.025: 0.955, 0.01: 1.035},
        'Exponential':       {0.15: 0.926, 0.10: 0.990, 0.05: 1.094,
                              0.025: 1.190, 0.01: 1.308},
        'Params_known':      {0.15: 1.138, 0.10: 1.224, 0.05: 1.358,
                              0.025: 1.480, 0.01: 1.628},
    }

    n      = len(data)
    sqrt_n = math.sqrt(n)
    D      = best['ks']

    if best_name == 'Normal':
        adj_stat = (sqrt_n - 0.01 + 0.85 / sqrt_n) * D
        c_crit   = ks_table['Normal_lilliefors'].get(alpha)
        variant  = 'Lilliefors-corrected (Normal with estimated mu, sigma)'
    elif best_name == 'Exponential':
        adj_stat = (sqrt_n + 0.26 + 0.5 / sqrt_n) * (D - 0.2 / n)
        c_crit   = ks_table['Exponential'].get(alpha)
        variant  = 'Adjusted (Exponential with estimated mean)'
    else:  # Uniform: a, b come from data min/max — treat as the generic row
        adj_stat = (sqrt_n + 0.12 + 0.11 / sqrt_n) * D
        c_crit   = ks_table['Params_known'].get(alpha)
        variant  = 'Generic KS critical value'

    if c_crit is None:
        print(f"  KS test: alpha={alpha} not in table "
              f"(supported: 0.15, 0.10, 0.05, 0.025, 0.01)")
    else:
        accepted = adj_stat < c_crit
        verdict  = 'ACCEPT H0' if accepted else 'REJECT H0'
        print(f"  KS test ({variant})")
        print(f"    alpha={alpha:.2f}  adj_stat={adj_stat:.4f}  "
              f"c_crit={c_crit:.4f}  -> {verdict}")

    # Return selected distribution name, parameters, and sampler function.
    return best_name, best['params'], best['sampler']


def fit_from_excel(xlsx_path: str) -> Dict[str, Callable]:
    """
    Load Excel sample data and fit distributions for each sheet.

    Returns a dict of callables:
        'friends_interarrival'  – inter-arrival time sampler (minutes)
        'main_stage_duration'   – show duration sampler (minutes)

    If the file cannot be read, falls back to sensible defaults.
    """

    # Load data from the Excel file.
    sheets = load_sample_data(xlsx_path)

    # Dictionary that will store the final sampler functions.
    samplers: Dict[str, Callable] = {}

    # ── Sheet 1: FriendsGroup inter-arrival times ─────────────────────────────

    # Get the first sheet name if it exists.
    sheet1_key = list(sheets.keys())[0] if sheets else None

    # If sheet 1 exists and has data, fit a distribution to it.
    if sheet1_key and sheets[sheet1_key]:
        data1 = sheets[sheet1_key]

        # Find the best distribution and get its sampler.
        _, _, sampler1 = best_fit(data1, label='FriendsGroup inter-arrival (min)')

        # Save the sampler.
        samplers['friends_interarrival'] = sampler1
    else:
        # If sheet 1 cannot be loaded, use a default sampler.
        print("[WARNING] Sheet 1 not loaded – using default Exponential(5 min) "
              "for FriendsGroup arrivals.")
        samplers['friends_interarrival'] = lambda: sample_exponential(5.0)

    # ── Sheet 2: MainStage show durations ─────────────────────────────────────

    # Get the second sheet name if it exists.
    sheet2_key = list(sheets.keys())[1] if len(sheets) > 1 else None

    # If sheet 2 exists and has data, fit a distribution to it.
    if sheet2_key and sheets[sheet2_key]:
        data2 = sheets[sheet2_key]

        # Find the best distribution and get its sampler.
        _, _, sampler2 = best_fit(data2, label='MainStage show duration (min)')

        # Save the sampler.
        samplers['main_stage_duration'] = sampler2
    else:
        # If sheet 2 cannot be loaded, use a default sampler.
        print("[WARNING] Sheet 2 not loaded – using default Exponential(60 min) "
              "for MainStage durations.")
        samplers['main_stage_duration'] = lambda: sample_exponential(60.0)

    # Return the ready-to-use samplers.
    return samplers